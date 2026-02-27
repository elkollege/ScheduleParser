import datetime
import itertools
import typing

import more_itertools
import openpyxl.cell.cell
import openpyxl.worksheet.worksheet

from . import constants
from . import models


# region Utils

def apply_substitutions_to_schedule(
        schedule: list[models.Period],
        substitutions: list[models.Substitution],
) -> list[models.Period]:
    for substitution in substitutions:
        try:
            schedule.remove(substitution.period)
        except ValueError:
            pass
        schedule.append(substitution.substitution)

    return sorted(
        [period for period in schedule if not period.is_empty],
        key=lambda period: (period.number, period.subgroup),
    )


def get_academic_week_number(date: datetime.datetime) -> int:
    current_week_number = date.isocalendar().week + 1

    first_academic_week_number = datetime.datetime(
        year=date.year - 1 if current_week_number < datetime.datetime(
            year=date.year - 1,
            month=9,
            day=1,
        ).isocalendar().week else date.year,
        month=9,
        day=1,
    ).isocalendar().week

    last_year_week_number = datetime.datetime(
        year=date.year - 1 if current_week_number < datetime.datetime(
            year=date.year - 1,
            month=12,
            day=28,
        ).isocalendar().week else date.year,
        month=12,
        day=28,
    ).isocalendar().week

    return current_week_number - first_academic_week_number + (
        last_year_week_number if current_week_number < first_academic_week_number else 0
    )


# endregion

# region Parsing

def parse_schedule(
        worksheet: openpyxl.worksheet.worksheet.Worksheet,
) -> typing.Generator[models.GroupSchedule]:
    def _get_periods(columns: tuple) -> typing.Generator[models.Period]:
        def _get_number(row: int) -> int:
            return int(
                worksheet.cell(
                    row=row,
                    column=constants.SCHEDULE_NUMBER_COL,
                ).value
            )

        def _get_subgroup(column: int) -> int:
            return int(
                worksheet.cell(
                    row=constants.SCHEDULE_SUBGROUP_ROW,
                    column=column,
                ).value
            )

        for period_columns in zip(
                *[
                    [
                        tuple(batched) for batched in itertools.batched(
                        column,
                        constants.SCHEDULE_PERIOD_HEIGHT,
                    )
                    ] for column in columns
                ]
        ):
            if isinstance(period_columns[1][0], openpyxl.cell.cell.MergedCell):
                yield models.Period(
                    number=_get_number(period_columns[0][0].row),
                    subgroup=constants.SCHEDULE_DEFAULT_SUBGROUP,
                    subject=str(period_columns[0][0].value),
                    lecturer=str(period_columns[0][1].value),
                    room=str(period_columns[3][0].value),
                )
            else:
                for first_col, second_col in [(0, 1), (2, 3)]:
                    if period_columns[first_col][0].value:
                        yield models.Period(
                            number=_get_number(period_columns[first_col][0].row),
                            subgroup=_get_subgroup(period_columns[first_col][0].column),
                            subject=str(period_columns[first_col][0].value),
                            lecturer=str(period_columns[first_col][1].value),
                            room=str(period_columns[second_col][0].value),
                        )

    def _get_day_schedules(columns: tuple) -> typing.Generator[models.DaySchedule]:
        def _get_weekday(row: int) -> int:
            return models.Weekday.from_string(
                string=str(
                    worksheet.cell(
                        row=row,
                        column=constants.SCHEDULE_WEEKDAY_COL,
                    ).value
                ),
            ).value

        for day_schedule_columns in zip(
                *[
                    tuple(
                        more_itertools.split_when(
                            column,
                            lambda
                                    cell,
                                    _
                            : cell.border.bottom.style == constants.SCHEDULE_WEEKDAY_BORDER,
                        )
                    ) for column in columns
                ]
        ):
            yield models.DaySchedule(
                weekday=_get_weekday(day_schedule_columns[0][0].row),
                periods_list=list(
                    _get_periods(day_schedule_columns)
                ),
            )

    def _get_group_name(column: int) -> str:
        return str(
            worksheet.cell(
                row=constants.SCHEDULE_GROUP_NAME_ROW,
                column=column,
            ).value
        )

    for group_schedule_columns in tuple(
            tuple(batched)[:-constants.SCHEDULE_JUNK_COLUMNS] for batched in itertools.batched(
                worksheet.iter_cols(
                    min_col=constants.SCHEDULE_MIN_COL,
                    min_row=constants.SCHEDULE_MIN_ROW,
                    max_row=worksheet.max_row - constants.SCHEDULE_BOTTOM_OFFSET,
                ),
                constants.SCHEDULE_GROUP_WIDTH,
            )
    ):
        yield models.GroupSchedule(
            group_name=_get_group_name(group_schedule_columns[0][0].column),
            day_schedules_list=list(
                _get_day_schedules(group_schedule_columns)
            ),
        )


def parse_substitutions(
        worksheet: openpyxl.worksheet.worksheet.Worksheet,
) -> typing.Generator[models.Substitution]:
    def _get_period(number: str, period_values: list[str]) -> models.Period:
        return models.Period(
            number=int(number),
            subgroup=int(period_values[0]),
            subject=str(period_values[1]),
            lecturer=str(period_values[2]),
            room=str(period_values[3]),
        )

    for substitution_row in worksheet.iter_rows(
            min_row=constants.SUBSTITUTIONS_MIN_ROW,
            min_col=constants.SUBSTITUTIONS_MIN_COL,
            max_col=constants.SUBSTITUTIONS_MAX_COL,
            values_only=True,
    ):
        group_name, period_number, *substitution_values = substitution_row

        yield models.Substitution(
            group_name=group_name,
            period=_get_period(period_number, substitution_values[:constants.SUBSTITUTION_WIDTH]),
            substitution=_get_period(period_number, substitution_values[constants.SUBSTITUTION_WIDTH:]),
        )

# endregion
