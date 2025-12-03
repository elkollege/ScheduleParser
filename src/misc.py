from __future__ import annotations

import itertools
import typing

import openpyxl.cell.cell
import openpyxl.worksheet.worksheet

import constants
import models


def parse_schedule(
        worksheet: openpyxl.worksheet.worksheet.Worksheet,
) -> typing.Generator[models.GroupScheduleContainer]:
    def _get_periods(
            columns: list[list[openpyxl.cell.cell.Cell]],
    ) -> typing.Generator[models.PeriodModel]:
        def _get_number(row: int) -> int:
            return int(
                worksheet.cell(
                    row=row,
                    column=constants.SCHEDULE_DATA_NUMBER_COLUMN,
                ).value
            )

        def _get_subgroup(column: int) -> int:
            return worksheet.cell(
                row=constants.SCHEDULE_DATA_SUBGROUP_ROW,
                column=column,
            ).value

        split_columns = [
            list(
                map(
                    lambda batched: list(batched),
                    itertools.batched(
                        column,
                        constants.SCHEDULE_PERIOD_HEIGHT,
                    ),
                )
            ) for column in columns
        ]

        for period in [list(current_columns) for current_columns in zip(*split_columns)]:
            if isinstance(period[1][0], openpyxl.cell.cell.MergedCell):
                yield models.PeriodModel(
                    data={
                        "lecturer": period[0][1].value,
                        "number": _get_number(period[0][0].row),
                        "room": period[3][0].value,
                        "subgroup": constants.SCHEDULE_DATA_NO_SUBGROUP,
                        "subject": period[0][0].value,
                    },
                )
            else:
                for first_column, second_column in [(0, 1), (2, 3)]:
                    if period[first_column][0].value:
                        yield models.PeriodModel(
                            data={
                                "lecturer": period[first_column][1].value,
                                "number": _get_number(period[first_column][0].row),
                                "room": period[second_column][0].value,
                                "subgroup": _get_subgroup(period[first_column][0].column),
                                "subject": period[first_column][1].value,
                            },
                        )

    def _get_day_schedule(
            columns: list[list[openpyxl.cell.cell.Cell]],
    ) -> typing.Generator[models.DayScheduleContainer]:
        def _get_weekday(row: int) -> models.Weekday:
            return models.Weekday.from_string(
                string=worksheet.cell(
                    row=row,
                    column=constants.SCHEDULE_DATA_WEEKDAY_COLUMN,
                ).value,
            )

        split_columns = []

        for column in columns:
            current_columns = []
            current_column = []

            for cell in column:
                current_column.append(cell)
                if cell.border.bottom.style == "medium":
                    current_columns.append(current_column)
                    current_column = []

            split_columns.append(current_columns)

        for day_schedule in [list(current_columns) for current_columns in zip(*split_columns)]:
            yield models.DayScheduleContainer(
                data={
                    "weekday": _get_weekday(day_schedule[0][0].row).value,
                    "schedule": list(
                        map(
                            lambda period: period._data,
                            _get_periods(day_schedule),
                        )
                    ),
                },
            )

    def _get_week_schedule(
            columns: list[list[openpyxl.cell.cell.Cell]],
    ) -> typing.Generator[models.WeekScheduleContainer]:
        def _check_cell_color(cell_index: int) -> bool:
            return columns[0][cell_index].fill.fgColor.rgb == constants.SCHEDULE_DATA_ODD_COLOR

        week_variants = []

        for parity in (True, False):
            week_variants.append(
                (
                    [
                        list(
                            filter(
                                lambda cell: _check_cell_color(column.index(cell)) == parity,
                                column,
                            )
                        ) for column in columns
                    ],
                    not parity,
                )
            )

        for week_variant in week_variants:
            yield models.WeekScheduleContainer(
                data={
                    "parity": week_variant[1],
                    "schedule": list(
                        map(
                            lambda day_schedule: day_schedule._data,
                            _get_day_schedule(week_variant[0]),
                        )
                    ),
                },
            )

    def _get_group(column: int) -> str:
        return worksheet.cell(
            row=constants.SCHEDULE_DATA_GROUP_ROW,
            column=column,
        ).value

    for group_columns in itertools.batched(
            list(worksheet.columns)[constants.SCHEDULE_SIDEBAR_WIDTH:constants.SCHEDULE_JUNK_OFFSET],
            constants.SCHEDULE_GROUP_WIDTH,
    ):
        group_columns = list(
            map(
                lambda column: column[constants.SCHEDULE_HEADER_INDEX:constants.SCHEDULE_FOOTER_OFFSET],
                list(group_columns[-constants.SCHEDULE_JUNK_OFFSET:]),
            )
        )

        yield models.GroupScheduleContainer(
            data={
                "group": _get_group(group_columns[0][0].column),
                "schedule": list(
                    map(
                        lambda week_schedule: week_schedule._data,
                        _get_week_schedule(group_columns),
                    )
                ),
            }
        )


def parse_substitutions(
        worksheet: openpyxl.worksheet.worksheet.Worksheet,
) -> typing.Generator[models.SubstitutionModel]:
    def _get_period(*args, number: int) -> models.PeriodModel:
        return models.PeriodModel(
            data={
                "lecturer": args[2],
                "number": number,
                "room": args[3],
                "subgroup": int(args[0]) if args[0] else 0,
                "subject": args[1],
            },
        )

    for row in list(worksheet.rows)[constants.SUBSTITUTIONS_HEADER_INDEX:]:
        row = list(
            map(
                lambda cell: cell.value,
                list(row[:constants.SUBSTITUTIONS_WIDTH]),
            )
        )

        if set(row) == {None}:
            continue

        group, period_number, *row_values = row
        period_number = int(period_number)

        yield models.SubstitutionModel(
            data={
                "group": group,
                "period": _get_period(
                    *row_values[4:],
                    number=period_number,
                )._data,
                "substitution": _get_period(
                    *row_values[:4],
                    number=period_number,
                )._data,
            },
        )
