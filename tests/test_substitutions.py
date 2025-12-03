from __future__ import annotations

import openpyxl
import pyquoks

import misc
import models


class TestSubstitutions(pyquoks.test.TestCase):
    _MODULE_NAME = __name__

    def test_parse_substitutions(self):
        workbook = openpyxl.load_workbook(
            filename=pyquoks.utils.get_path("resources/tables/substitutions.xlsx"),
        )

        worksheet = workbook.worksheets[0]

        for substitution in list(misc.parse_substitutions(worksheet)):
            self.assert_type(
                func_name=self.test_parse_substitutions.__name__,
                test_data=substitution,
                test_type=models.SubstitutionModel,
            )

    def test_parse_substitutions_incorrect(self):
        workbook = openpyxl.load_workbook(
            filename=pyquoks.utils.get_path("resources/tables/substitutions_incorrect.xlsx"),
        )

        worksheet = workbook.worksheets[0]

        for substitution in list(misc.parse_substitutions(worksheet)):
            self.assert_type(
                func_name=self.test_parse_substitutions_incorrect.__name__,
                test_data=substitution,
                test_type=models.SubstitutionModel,
            )

    def test_parse_substitutions_old(self):
        workbook = openpyxl.load_workbook(
            filename=pyquoks.utils.get_path("resources/tables/substitutions_old.xlsx"),
        )

        worksheet = workbook.worksheets[0]

        for substitution in list(misc.parse_substitutions(worksheet)):
            self.assert_type(
                func_name=self.test_parse_substitutions_old.__name__,
                test_data=substitution,
                test_type=models.SubstitutionModel,
            )
