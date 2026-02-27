import datetime

import openpyxl
import pyquoks

import schedule_parser.models
import schedule_parser.utils
from . import models


class TestUtils:
    @classmethod
    def setup_class(cls) -> None:
        cls._TEST_SUBSTITUTIONS = [
            models.TestSubstitution(
                date=datetime.date(2026, 2, 11),
                test_schedules_list=[
                    models.TestCorrectSchedule(
                        group_name="ЭОЭЭО 25-02",
                        correct_schedule="""\
                            1. Информ. (теор.) | 305
                            2. Физика | 110
                            3. Математика | 411
                            4. Русс.яз. | 213
                            5. Биология | 410
                        """,
                    ),
                ],
            ),
            models.TestSubstitution(
                date=datetime.date(2026, 2, 12),
                test_schedules_list=[
                    models.TestCorrectSchedule(
                        group_name="ЭОЭЭО 25-02",
                        correct_schedule="""\
                            0. Классный час | 410
                            1. Химия | 102
                            2. Физ-ра | с.з.
                            3. Лит-ра | 213
                            4. История | 301
                        """,
                    ),
                ],
            ),
            models.TestSubstitution(
                date=datetime.date(2026, 2, 13),
                test_schedules_list=[
                    models.TestCorrectSchedule(
                        group_name="ЭОЭЭО 25-02",
                        correct_schedule="""\
                            1. (1) Информ. (пр.) | 305
                            2. ОБиЗР | 201
                            3. Математика | 308
                            4. Инж. граф. | 210
                            5. История | 301
                        """,
                    ),
                ],
            ),
            models.TestSubstitution(
                date=datetime.date(2026, 2, 16),
                test_schedules_list=[
                    models.TestCorrectSchedule(
                        group_name="ЭОЭЭО 25-02",
                        correct_schedule="""\
                            4. География | 401
                            5. Физика | 110
                        """,
                    ),
                ],
            ),
            models.TestSubstitution(
                date=datetime.date(2026, 2, 17),
                test_schedules_list=[
                    models.TestCorrectSchedule(
                        group_name="ЭОЭЭО 25-02",
                        correct_schedule="""\
                            1. Лит-ра | 213
                            2. География | 401
                            3. Математика | 411
                            4. Химия | 103М
                        """,
                    ),
                ],
            ),
            models.TestSubstitution(
                date=datetime.date(2026, 2, 18),
                test_schedules_list=[
                    models.TestCorrectSchedule(
                        group_name="ЭОЭЭО 25-02",
                        correct_schedule="""\
                            1. Математика | 308
                            2. Физика | 110
                            3. Математика | 308
                            4. Русс.яз. | 213
                            5. Биология | 410
                        """,
                    ),
                ],
            ),
            models.TestSubstitution(
                date=datetime.date(2026, 2, 19),
                test_schedules_list=[
                    models.TestCorrectSchedule(
                        group_name="ЭОЭЭО 25-02",
                        correct_schedule="""\
                            0. Классный час | 410
                            1. Химия | 102
                            2. Физ-ра | с.з.
                            3. Биология | 410
                            4. История | 301
                        """,
                    ),
                ],
            ),
            models.TestSubstitution(
                date=datetime.date(2026, 2, 20),
                test_schedules_list=[
                    models.TestCorrectSchedule(
                        group_name="ЭОЭЭО 25-02",
                        correct_schedule="""\
                            0. Разговоры о важном | 410
                            1. (1) Информ. (пр.) | 305
                            1. (2) Ин. язык | 309
                            2. ОБиЗР | 201
                            3. Физ-ра | спортзал
                            4. (1) Ин. язык | 309
                            4. (2) Информ. (пр.) | 305
                        """,
                    ),
                ],
            ),
        ]

    def test_parse_schedule(self):
        workbook = openpyxl.load_workbook(
            filename=pyquoks.utils.get_path("resources/tables/schedule.xlsx"),
        )

        for model in schedule_parser.utils.parse_schedule(workbook.worksheets[0]):
            assert isinstance(model, schedule_parser.models.GroupSchedule), "objects in parsed schedules list"

    def test_parse_substitutions(self):
        for test_substitution in self._TEST_SUBSTITUTIONS:
            string_date = test_substitution.date.strftime("%d_%m_%y")

            workbook = openpyxl.load_workbook(
                filename=pyquoks.utils.get_path(f"resources/tables/substitutions_{string_date}.xlsx"),
            )

            for substitution in schedule_parser.utils.parse_substitutions(workbook.worksheets[0]):
                assert isinstance(
                    substitution,
                    schedule_parser.models.Substitution,
                ), f"({test_substitution.date}) objects in parsed substitutions list"

    def test_apply_substitutions_to_schedule(self):
        workbook_schedule = openpyxl.load_workbook(
            filename=pyquoks.utils.get_path("resources/tables/schedule.xlsx"),
        )

        for test_substitution in self._TEST_SUBSTITUTIONS:
            for test_schedule in test_substitution.test_schedules_list:
                string_date = test_substitution.date.strftime("%d_%m_%y")

                workbook_substitutions = openpyxl.load_workbook(
                    filename=pyquoks.utils.get_path(f"resources/tables/substitutions_{string_date}.xlsx"),
                )

                current_schedule = schedule_parser.utils.apply_substitutions_to_schedule(
                    schedule=schedule_parser.models.GroupSchedule.get_group_schedule_by_group_name(
                        iterable=list(
                            schedule_parser.utils.parse_schedule(
                                worksheet=workbook_schedule.worksheets[0],
                            )
                        ),
                        group_name=test_schedule.group_name,
                    ).get_day_schedule_by_weekday(
                        weekday=test_substitution.date.weekday(),
                    ).periods_list,
                    substitutions=schedule_parser.models.Substitution.get_substitutions_by_group_name(
                        iterable=list(
                            schedule_parser.utils.parse_substitutions(
                                worksheet=workbook_substitutions.worksheets[0],
                            )
                        ),
                        group_name=test_schedule.group_name,
                    ),
                )

                assert "\n".join(
                    period.readable for period in current_schedule
                ) == test_schedule.correct_schedule_dedent, f"({test_substitution.date}) compare parsed schedule with correct one"
